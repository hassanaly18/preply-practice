from openpyxl import Workbook

wb = Workbook()

sheet = wb.active()

sheet["A1"] = "Product"
sheet["B1"] = "Price"

wb.save("menu.xlsx")












# file = openpyxl.load_workbook("Book1.xlsx")

# sheet = file["Sheet1"]

# sheet["A5"] = "Harry"
# file.save("Book1.xlsx")


# marks = sheet["B2"].value
# print(marks)

# for row in sheet.iter_rows(values_only=True):
#     print(row)